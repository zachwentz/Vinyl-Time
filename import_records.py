
def import_rec():

    import json
    import openpyxl 

    #path = "C:\\Users\\zwentz\\records.xlsx"
    path = "records1.xlsx"
    rec = openpyxl.load_workbook(path) 
    rec_sheet = rec.active
    max_col = rec_sheet.max_column
    max_row = rec_sheet.max_row
    rec_val = rec_sheet['A1':'D' + str(max_row)]

    records = {}
    
    for c1, c2, c3, c4, c5, c6 in rec_val:
        if c1.value in records.keys():
            records[c1.value].update({c2.value:{'Sides':c3.value, 'Color':c4.value,'Speed':c5.value,'Size':c6.value}})
        else:
            records[c1.value] = {c2.value:{'Sides':c3.value,'Color':c4.value,'Speed':c5.value,'Size':c6.value}}

    with open('record_collection.json','w') as outfile:
        json.dump(records,outfile)